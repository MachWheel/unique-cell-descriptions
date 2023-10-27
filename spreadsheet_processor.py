import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook


def process_spreadsheet(input_file_path):
    # Read the spreadsheet
    df = pd.read_excel(input_file_path)

    # Remove the character ";" from the entire DataFrame
    df = df.replace(';', '', regex=True)

    # Extract unique values from column F ("Description"), sort them, and create a new DataFrame
    new_df = pd.DataFrame(df['Description'].unique(), columns=['Description']).sort_values(by='Description')

    return new_df


def save_spreadsheet(df, output_file_path):
    # Create a new workbook and select the active worksheet
    book = Workbook()
    sheet = book.active
    sheet.title = "Sheet1"

    # Create styles for the header
    bold_font = Font(bold=True)
    center_aligned_text = Alignment(horizontal="center")

    # Start writing the DataFrame to the open workbook
    # First, let's handle the header row separately to apply styles
    for c_idx, col_name in enumerate(df.columns, 1):
        cell = sheet.cell(row=1, column=c_idx, value=col_name)
        cell.font = bold_font
        cell.alignment = center_aligned_text

    # Now, write the actual data from the DataFrame
    for r_idx, row in enumerate(df.values, 2):  # We start from the 2nd row as the 1st row is the header
        for c_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)

    # Adjust the width of the column (we don't divide by 7 in this case, to ensure we get a width of 200 units)
    col_letter = get_column_letter(1)
    sheet.column_dimensions[col_letter].width = 200

    # Save the workbook to the specified file path
    book.save(output_file_path)
